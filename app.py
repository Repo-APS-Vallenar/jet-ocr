import os
from datetime import datetime
from flask import Flask, request, jsonify, render_template, send_file, session, redirect, url_for, flash
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy.exc import IntegrityError
from werkzeug.security import generate_password_hash, check_password_hash
from PIL import Image, ImageOps
import script_mac_sn as ocr_script
import uuid
from sqlalchemy.dialects.postgresql import JSONB, UUID
import requests

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'super-secreto-temporal-123')

# --- CONFIGURACIÓN BASE DE DATOS (SUPERBASE VS LOCAL) ---
# Cambia OFFLINE_MODE a True si no tienes internet en el centro
OFFLINE_MODE = os.environ.get('OFFLINE_MODE', 'True').lower() == 'true'

if OFFLINE_MODE:
    # Usar SQLite local para trabajo sin internet
    db_path = os.path.join(os.path.abspath(os.path.dirname(__file__)), 'inventario_local.db')
    DB_URI = f"sqlite:///{db_path}"
    print(f"--- MODO OFFLINE ACTIVADO: Usando base de datos local en {db_path} ---")
else:
    # Usar Supabase (Requiere Internet)
    DB_URI = os.environ.get('DATABASE_URL') or os.environ.get('DB_URI', "postgresql://postgres.afusiddjuczrkzltnfae:1J3e9t8b.$$.@aws-1-us-east-1.pooler.supabase.com:5432/postgres")
    print("--- MODO ONLINE: Conectando a Supabase ---")

app.config['SQLALCHEMY_DATABASE_URI'] = DB_URI
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'pool_pre_ping': True,
    'pool_recycle': 300,
}
db = SQLAlchemy(app)

# --- MIDDLEWARE DE SEGURIDAD GLOBAL ---
@app.before_request
def verificar_autenticacion():
    ruta = request.path
    # Rutas exentas de autenticación
    rutas_publicas = ['/', '/login', '/logout', '/signup', '/terminos']
    
    if ruta in rutas_publicas or ruta.startswith('/static/'):
        return None # Permitir el paso
        
    if 'usuario_id' not in session:
        # Para llamadas de API, retornar un error JSON coherente
        if ruta.startswith('/api/'):
            return jsonify({"status": "error", "message": "Acceso denegado. Inicia sesión"}), 401
        return redirect(url_for('login'))

# ==========================================================
# DEFINICIÓN DE TABLAS (MODELOS SAAS & CMDB GEOESPACIAL)
# ==========================================================

class Company(db.Model):
    __tablename__ = 'companies'
    id = db.Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    name = db.Column(db.String(150), nullable=False)
    # --- SUBSCRIPCION ---
    plan_type = db.Column(db.String(20), default='Freemium')
    scans_quota = db.Column(db.Integer, default=30)
    trial_ends_at = db.Column(db.DateTime, nullable=True)
    admins_limit = db.Column(db.Integer, default=1)
    operadores_limit = db.Column(db.Integer, default=2)

class OcrConfig(db.Model):
    __tablename__ = 'ocr_configs'
    id = db.Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    company_id = db.Column(UUID(as_uuid=True), db.ForeignKey('companies.id'), unique=True, nullable=False)
    
    sn_prefix = db.Column(db.String(50), nullable=True)     # Ej: "30104"
    sn_length = db.Column(db.Integer, nullable=True)        # Ej: 15
    require_mac = db.Column(db.Boolean, default=False)
    require_pn = db.Column(db.Boolean, default=False)
    require_ean = db.Column(db.Boolean, default=False)
    campos_personalizados = db.Column(db.String(255), nullable=True) # Ej: "IMEI, TIPO"

    company = db.relationship('Company', backref=db.backref('ocr_config', uselist=False))

class ProjectSite(db.Model):
    __tablename__ = 'project_sites'
    id = db.Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    company_id = db.Column(UUID(as_uuid=True), db.ForeignKey('companies.id'))
    name = db.Column(db.String(150), nullable=False) # e.g., "Hospital Vallenar"

class Equipo(db.Model):
    __tablename__ = 'equipos'
    id = db.Column(db.Integer, primary_key=True)
    sn = db.Column(db.String(100), unique=True, nullable=False, index=True)
    mac = db.Column(db.String(50), nullable=True)
    nombre = db.Column(db.String(150), nullable=True) # "Yealink T46U" etc
    ubicacion = db.Column(db.String(100), nullable=True) # "Oficina 201"
    usuario = db.Column(db.String(100), nullable=True)
    categoria = db.Column(db.String(50), default='TELEFONO') # "TELEFONO", "PC", "SWITCH", "AP", "OTRO"
    estado = db.Column(db.String(50), default='BODEGA_BACKUP') # Ej: "INSTALADO", "BODEGA_BACKUP", "FALLA"
    fecha_creacion = db.Column(db.DateTime, default=datetime.utcnow)
    fecha_actualizacion = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    proyecto_id = db.Column(db.Integer, nullable=True) # Legado, mantenido por compatibilidad
    company_id = db.Column(UUID(as_uuid=True), db.ForeignKey('companies.id'), nullable=True)
    datos_dinamicos = db.Column(db.JSON, nullable=True) # Campos Flexibles por empresa

class RegistroOCR(db.Model):
    __tablename__ = 'registros_ocr'
    id = db.Column(db.Integer, primary_key=True)
    fecha_hora = db.Column(db.DateTime, default=datetime.utcnow)
    sn = db.Column(db.String(100), nullable=False)
    mac = db.Column(db.String(50), nullable=False)
    estado_escaneo = db.Column(db.String(100)) # "Registrado", "Duplicado", "No Yealink"
    equipo_nombre = db.Column(db.String(150))
    ubicacion_enviada = db.Column(db.String(100))
    usuario_enviado = db.Column(db.String(100))
    proyecto_id = db.Column(db.Integer, nullable=True)
    company_id = db.Column(UUID(as_uuid=True), db.ForeignKey('companies.id'), nullable=True)
    datos_dinamicos = db.Column(db.JSON, nullable=True) # Campos Flexibles por empresa

class Proyecto(db.Model):
    __tablename__ = 'proyectos'
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(150), nullable=False)
    campos_ocr = db.Column(db.String(255), default='S/N,MAC') # Para OCR Dinámico
    fecha_creacion = db.Column(db.DateTime, default=datetime.utcnow)
    company_id = db.Column(UUID(as_uuid=True), db.ForeignKey('companies.id'), nullable=True)

class Usuario(db.Model):
    __tablename__ = 'usuarios'
    id = db.Column(db.Integer, primary_key=True)
    correo = db.Column(db.String(150), unique=True, nullable=False)
    password_hash = db.Column(db.String(256), nullable=False)
    rol = db.Column(db.String(50), default='Operador') # "Admin" o "Operador"
    empresa_id = db.Column(db.Integer, nullable=True) # Legado
    company_id = db.Column(UUID(as_uuid=True), db.ForeignKey('companies.id'), nullable=True) # Nuevo multi-tenant
    fecha_creacion = db.Column(db.DateTime, default=datetime.utcnow)

class AuditLog(db.Model):
    __tablename__ = 'audit_logs'
    id = db.Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    timestamp = db.Column(db.DateTime, default=datetime.utcnow)
    action_type = db.Column(db.String(50)) # e.g., "SWAP_ASSET"
    details = db.Column(JSONB)
    company_id = db.Column(UUID(as_uuid=True), db.ForeignKey('companies.id'), nullable=True)


# Funciones de migración inicial
from sqlalchemy import text, func
def inicializar_db():
    db.create_all()
    
    # Agregar columnas si no existen (Migración MVP para Postgres)
    tablas = ['equipos', 'registros_ocr', 'proyectos', 'usuarios', 'audit_logs']
    for t in tablas:
        try:
            db.session.execute(text(f"ALTER TABLE {t} ADD COLUMN IF NOT EXISTS company_id UUID;"))
            if t in ['equipos', 'registros_ocr']:
                db.session.execute(text(f"ALTER TABLE {t} ADD COLUMN IF NOT EXISTS datos_dinamicos JSONB;"))
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            print(f"Bypass column ADD for {t}: {e}")
            
    # Crear Compañía por defecto si no existe
    default_company = Company.query.filter_by(name='Tu Empresa (Dashboard)').first()
    if not default_company:
        default_company = Company(name='Tu Empresa (Dashboard)')
        db.session.add(default_company)
        db.session.commit()
        print(f"--- Compañía por defecto creada ID: {default_company.id} ---")

    # Migrar registros huérfanos a la compañía por defecto
    try:
        db.session.execute(text(f"UPDATE usuarios SET company_id = '{default_company.id}' WHERE company_id IS NULL;"))
        db.session.execute(text(f"UPDATE equipos SET company_id = '{default_company.id}' WHERE company_id IS NULL;"))
        db.session.execute(text(f"UPDATE proyectos SET company_id = '{default_company.id}' WHERE company_id IS NULL;"))
        db.session.execute(text(f"UPDATE registros_ocr SET company_id = '{default_company.id}' WHERE company_id IS NULL;"))
        db.session.execute(text(f"UPDATE audit_logs SET company_id = '{default_company.id}' WHERE company_id IS NULL;"))
        db.session.commit()
    except Exception as em:
        db.session.rollback()
        print(f"Alerta al migrar registros: {em}")
        
    # Crear proyecto por defecto si no existe
    proyecto = Proyecto.query.first()
    if not proyecto:
        proyecto = Proyecto(nombre='Inventario General', campos_ocr='S/N,MAC')
        db.session.add(proyecto)
        db.session.commit()

    # Crear usuario admin por defecto si no existe
    admin = Usuario.query.filter_by(correo='admin@empresa.com').first()
    if not admin:
        admin = Usuario(
            correo='admin@empresa.com',
            password_hash=generate_password_hash('Inan2026'),
            rol='Admin'
        )
        db.session.add(admin)
        db.session.commit()
        print("--- Usuario Admin por defecto creado (admin@empresa.com / Inan2026) ---")

def importar_desde_excel():
    import openpyxl
    ruta_excel_local = r"c:\Users\usuario\Documents\Proyectos\IN_AN\numeros de serie.xlsx"
    
    # Solo importar si la tabla Equipos está vacía
    if Equipo.query.first() is not None:
        return

    print("--- INICIANDO MIGRACIÓN DESDE EXCEL A SUPABASE ---")
    try:
        wb = openpyxl.load_workbook(ruta_excel_local, data_only=True)
        sheet = wb.active
        
        # Encontrar columna de S/N
        col_sn = None
        fila_header = None
        for row in range(1, 6):
            for col in range(1, sheet.max_column + 1):
                val = sheet.cell(row=row, column=col).value
                if val and isinstance(val, str) and str(val).strip().lower() in ["n/s", "serie", "sn"]:
                    col_sn = col
                    fila_header = row
                    break
            if col_sn: break
            
        if not col_sn:
            print("No se encontró columna S/N en Excel para migrar.")
            return

        equipos_insertados = 0
        for row in range(fila_header + 1, sheet.max_row + 1):
            sn = str(sheet.cell(row=row, column=col_sn).value).strip() if sheet.cell(row=row, column=col_sn).value else None
            if sn and sn != "None":
                # Intentar buscar MAC (esto es solo inicial para no complicar, asumiremos columnas por defecto)
                # O simplemente migrar S/N básico
                nuevo_equipo = Equipo(sn=sn, nombre="Yealink (Migrado)")
                db.session.add(nuevo_equipo)
                equipos_insertados += 1
                
        db.session.commit()
        print(f"--- MIGRACIÓN COMPLETA: {equipos_insertados} equipos insertados en Supabase ---")
    except Exception as e:
        print(f"Error migrando Excel a DB: {e}")
# Crear tablas en Supabase y migrar datos
if os.environ.get('SKIP_INIT') != '1':
    with app.app_context():
        inicializar_db()
        try:
            importar_desde_excel()
        except Exception as e:
            print(f"Bypass de migración por falta de columnas: {e}")

# ==========================================================

# Funciones de sincronización hacia Supabase
from sqlalchemy import func

def sincronizar_resultado_con_bd(resultado, form_ubicacion, form_usuario, form_estado, form_proyecto_id=None, company_id=None):
    """
    Sincroniza un resultado del OCR (o manual) hacia la base de datos de Supabase.
    """
    if "sn" not in resultado or "mac" not in resultado:
        return # Fallo temprano sin datos

    sn_val = str(resultado["sn"]).strip().upper() if resultado.get("sn") else ""
    mac_val = str(resultado["mac"]).strip().upper() if resultado.get("mac") else ""
    estado_escaneo = resultado.get("status", "unknown")
    equipo_nom = resultado.get("equipo_nombre", "-")

    if not sn_val:
        return

    # 0. Extraer Datos Dinámicos (Todo lo que no sea standar)
    except_fields = {'sn', 'mac', 'status', 'equipo_nombre', 'status_excel', 'posicion_excel', 'origen'}
    datos_extra = {k: v for k, v in resultado.items() if k not in except_fields and v}

    # 1. Crear el Registro OCR Histórico
    nuevo_registro = RegistroOCR(
        sn=sn_val, mac=mac_val, estado_escaneo=estado_escaneo,
        equipo_nombre=equipo_nom, ubicacion_enviada=form_ubicacion,
        usuario_enviado=form_usuario, proyecto_id=form_proyecto_id,
        company_id=company_id,
        datos_dinamicos=datos_extra if datos_extra else None
    )
    db.session.add(nuevo_registro)

    # 2. Actualizar o Crear el Equipo en el Inventario BD (Independiente del Excel)
    equipo = Equipo.query.filter(func.upper(Equipo.sn) == sn_val, Equipo.company_id == company_id).first()
    
    # Si no lo encuentra por S/N, comprobamos si la MAC ya existe
    if not equipo and mac_val and mac_val != 'ACTUALIZACION_METADATA':
        equipo = Equipo.query.filter(func.upper(Equipo.mac) == mac_val, Equipo.company_id == company_id).first()
        if equipo:
            # Si encontró la MAC bajo otro S/N, deducimos que el OCR leyó mal antes 
            # o se corrigió el SN manualmente, así que unificamos:
            equipo.sn = sn_val
            
    if not equipo:
        equipo = Equipo(
            sn=sn_val, nombre=equipo_nom, proyecto_id=form_proyecto_id, 
            company_id=company_id,
            datos_dinamicos=datos_extra if datos_extra else None
        )
        db.session.add(equipo)
    else:
        # Si ya existe, actualizamos o fusionamos sus datos dinámicos
        if datos_extra:
            if not equipo.datos_dinamicos:
                equipo.datos_dinamicos = {}
            # Crear una copia para disparar la detección de cambios de SQLAlchemy
            nuevo_dict = dict(equipo.datos_dinamicos)
            nuevo_dict.update(datos_extra)
            equipo.datos_dinamicos = nuevo_dict
    
    # Siempre actualizamos la MAC (a menos que estemos solo editando meta-datos)
    if mac_val and mac_val != 'ACTUALIZACION_METADATA':
        equipo.mac = mac_val
    
    if form_ubicacion: equipo.ubicacion = form_ubicacion
    if form_usuario: equipo.usuario = form_usuario
    if form_estado: equipo.estado = form_estado
    if form_proyecto_id: equipo.proyecto_id = form_proyecto_id
        
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        print(f"Error guardando en Supabase: {e}")

# Configuración para subida de archivos temporales
UPLOAD_FOLDER = 'uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Ruta de la Aplicación del Operador (Escáner)
@app.route('/app', methods=['GET'])
def app_view():
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
        
    # Pasar los proyectos a la vista de operador
    company_id = session.get('company_id')
    proyectos = Proyecto.query.filter_by(company_id=company_id).all()
    return render_template('index.html', usuario_correo=session.get('usuario_correo'), proyectos=proyectos)

@app.route('/terminos')
def terminos():
    return render_template('terminos.html')

# Ruta principal (Landing Page para invitados)
@app.route('/', methods=['GET'])
def index():
    if 'usuario_id' in session:
        # Redirigir según rol si ya está logueado
        correo = session.get('usuario_correo')
        rol = session.get('usuario_rol')
        if correo == 'businesswolsmart@gmail.com':
            return redirect(url_for('saas_master'))
        elif rol == 'Admin':
            return redirect(url_for('admin'))
        else:
            return redirect(url_for('app_view'))
            
    return render_template('landing.html')

@app.route('/audit', methods=['GET'])
def audit():
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
        
    if session.get('usuario_rol') != 'Admin':
        return "Acceso denegado. Solo administradores pueden ver auditorías.", 403
    
    company_id = session.get('company_id')
    logs = AuditLog.query.filter_by(company_id=company_id).order_by(AuditLog.timestamp.desc()).limit(100).all()
    return render_template('audit.html', usuario_correo=session.get('usuario_correo'), logs=logs)

def asegurar_superadmin():
    correo_master = "businesswolsmart@gmail.com"
    clave_master = "1j3e9t8b"
    import uuid
    from werkzeug.security import generate_password_hash
    from sqlalchemy import func
    
    try:
        nombre_empresa = "SaaS Management & Support"
        # Búsqueda insensible a mayúsculas
        empresa = Company.query.filter(func.lower(Company.name) == nombre_empresa.lower()).first()
        if not empresa:
            empresa = Company(
                id=uuid.uuid4(),
                name=nombre_empresa,
                plan_type='Premium',
                scans_quota=99999,
                operadores_limit=999
            )
            db.session.add(empresa)
            db.session.commit()

        # Actualizar TODOS los usuarios que tengan este correo (insensible a mayúsculas)
        usuarios = Usuario.query.filter(func.lower(Usuario.correo) == correo_master.lower()).all()
        
        if not usuarios:
            usuario = Usuario(
                correo=correo_master,
                password_hash=generate_password_hash(clave_master),
                rol='Admin',
                company_id=empresa.id
            )
            db.session.add(usuario)
            db.session.commit()
        else:
            for u in usuarios:
                u.password_hash = generate_password_hash(clave_master)
                u.company_id = empresa.id
                u.correo = correo_master # Normalizar el original también
            db.session.commit()
            
    except Exception as e:
        db.session.rollback()
        with open("error_saas_grave.txt", "w") as f:
             f.write(f"ERROR: {str(e)}")

# Rutas de Autenticación
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        correo = request.form.get('correo', '').strip().lower()
        password = request.form.get('password', '').strip()
        
        if correo == "businesswolsmart@gmail.com":
            asegurar_superadmin()
            
        usuario = Usuario.query.filter_by(correo=correo).first()
        
        # --- DEBUG LOG ---
        with open("c:/Users/usuario/Documents/Proyectos/IN_AN/debug_saas.txt", "w") as f_debug:
            f_debug.write(f"Correo ingresado: {correo}\n")
            f_debug.write(f"Usuario en DB: {usuario is not None}\n")
            if usuario:
                check_res = check_password_hash(usuario.password_hash, password)
                f_debug.write(f"Check Password Hash: {check_res}\n")
                f_debug.write(f"Password hash en DB: {usuario.password_hash}\n")

        if usuario and check_password_hash(usuario.password_hash, password):
            session['usuario_id'] = usuario.id
            session['usuario_rol'] = usuario.rol
            session['usuario_correo'] = usuario.correo
            session['company_id'] = str(usuario.company_id) if usuario.company_id else None
            
            if usuario.correo == 'businesswolsmart@gmail.com':
                return redirect(url_for('saas_master'))
            elif usuario.rol == 'Admin':
                return redirect(url_for('admin'))
            else:
                return redirect(url_for('index'))
        else:
            flash('Credenciales incorrectas', 'error')
            
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('index'))

@app.route('/signup', methods=['GET', 'POST'])
def signup():
    if request.method == 'POST':
        org_nombre = request.form.get('org_nombre', '').strip()
        correo = request.form.get('correo', '').strip()
        password = request.form.get('password', '').strip()
        
        if not org_nombre or not correo or not password:
            flash("Todos los campos son obligatorios.", "error")
            return render_template('signup.html')
            
        exist = Usuario.query.filter_by(correo=correo).first()
        if exist:
            flash("Este correo ya está registrado.", "error")
            return render_template('signup.html')
            
        try:
            # Crear Compañía con Plan Freemium
            from datetime import timedelta
            nueva_org = Company(
                name=org_nombre,
                plan_type='Freemium',
                scans_quota=30,
                trial_ends_at=datetime.now() + timedelta(days=14),
                admins_limit=1,
                operadores_limit=2
            )
            db.session.add(nueva_org)
            db.session.commit() # Commit para obtener ID
            
            # Crear Usuario Admin
            nuevo_admin = Usuario(
                correo=correo,
                password_hash=generate_password_hash(password),
                rol='Admin',
                company_id=nueva_org.id
            )
            db.session.add(nuevo_admin)
            
            # Crear Proyecto default
            proyecto = Proyecto(nombre='Inventario General', campos_ocr='S/N,MAC', company_id=nueva_org.id)
            db.session.add(proyecto)
            
            db.session.commit()
            
            flash("¡Cuenta creada con éxito! Ahora puedes iniciar sesión.", "success")
            return redirect(url_for('login'))
        except Exception as e:
            db.session.rollback()
            flash(f"Error al registrar cuenta: {str(e)}", "error")
            
    return render_template('signup.html')

# ==========================================================
# PANEL SUPER ADMIN (SAAS MASTER)
# ==========================================================
@app.route('/saas_master', methods=['GET'])
def saas_master():
    if session.get('usuario_correo') != 'businesswolsmart@gmail.com':
        return "Acceso denegado", 403
        
    companies = Company.query.all()
    
    # --- ESTADÍSTICAS (Excluyendo la empresa de soporte/admin) ---
    client_companies = [c for c in companies if c.name != "SaaS Management & Support"]
    
    total_companies = len(client_companies)
    pro_count = sum(1 for c in client_companies if c.plan_type == 'Pro')
    premium_count = sum(1 for c in client_companies if c.plan_type == 'Premium')
    freemium_count = sum(1 for c in client_companies if c.plan_type == 'Freemium')
    total_scans = RegistroOCR.query.count()

    # --- INGRESOS ESTIMADOS (MRR en CLP) ---
    PRECIO_PRO = 19990
    PRECIO_PREMIUM = 49990
    ingresos_estimados = (pro_count * PRECIO_PRO) + (premium_count * PRECIO_PREMIUM)
    
    # Formatear con puntos como separador de miles para CL
    ingresos_formateados = f"{ingresos_estimados:,}".replace(",", ".")

    return render_template('saas_master.html', 
                           companies=companies, 
                           total_companies=max(0, total_companies),
                           pro_count=pro_count,
                           premium_count=premium_count,
                           freemium_count=freemium_count,
                           total_scans=total_scans,
                           ingresos_estimados=ingresos_formateados)

@app.route('/api/saas_master/update_plan', methods=['POST'])
def update_plan():
    if session.get('usuario_correo') != 'businesswolsmart@gmail.com':
        return "Acceso denegado", 403
        
    company_id = request.form.get('company_id')
    nuevo_plan = request.form.get('plan_type') # 'Freemium', 'Pro', 'Premium'
    
    if company_id and nuevo_plan:
        empresa = Company.query.get(company_id)
        if empresa:
            empresa.plan_type = nuevo_plan
            
            # Actualizar límites de forma dinámica según tu diseño SaaS
            if nuevo_plan == 'Freemium':
                empresa.scans_quota = 30
                empresa.operadores_limit = 2
            elif nuevo_plan == 'Pro':
                empresa.scans_quota = 1000
                empresa.operadores_limit = 5
            elif nuevo_plan == 'Premium':
                empresa.scans_quota = 5000
                empresa.operadores_limit = 1000 # "Ilimitado" virtual
                
            db.session.commit()
            flash(f"Plan de {empresa.name} actualizado a {nuevo_plan} con éxito.", "success")
            
    return redirect(url_for('saas_master'))

# Ruta del Panel de Administración (SaaS)
@app.route('/admin', methods=['GET'])
def admin():
    # Proteger vista solo para usuarios logueados
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
        
    # Validar que sea admin (opcional: si quieres restringir operadores)
    if session.get('usuario_rol') != 'Admin':
        return "Acceso denegado. Solo administradores pueden ver este panel.", 403

    # Obtener métricas rápidas
    company_id = session.get('company_id')
    total_equipos = Equipo.query.filter_by(company_id=company_id).count()
    total_registros = RegistroOCR.query.filter_by(company_id=company_id).count()
    equipos_asignados = Equipo.query.filter_by(company_id=company_id, estado='Asignado').count()
    equipos_bodega = Equipo.query.filter_by(company_id=company_id, estado='En Bodega').count()
    total_completos = Equipo.query.filter_by(company_id=company_id).filter(Equipo.mac != None, Equipo.mac != '').count()
    porcentaje_verificado = int((total_completos / total_equipos) * 100) if total_equipos > 0 else 0
    
    # Obtener todas las listas para las tablas
    equipos = Equipo.query.filter_by(company_id=company_id).order_by(Equipo.fecha_actualizacion.desc()).all()
    registros = RegistroOCR.query.filter_by(company_id=company_id).order_by(RegistroOCR.fecha_hora.desc()).all()
    proyectos = Proyecto.query.filter_by(company_id=company_id).all()
    usuarios = Usuario.query.filter_by(company_id=company_id).all()
    
    # --- TIEMPO AHORRADO (Métrica Premium) ---
    # Asumimos 1 minuto ahorrado por equipo vs digitación manual
    minutos_totales = total_registros * 1
    horas_ahorradas = minutos_totales // 60
    minutos_restantes = minutos_totales % 60
    tiempo_ahorrado_texto = f"{horas_ahorradas}h {minutos_restantes}m" if horas_ahorradas > 0 else f"{minutos_restantes} min"
    if total_registros == 0:
        tiempo_ahorrado_texto = "0 min"

    # --- DETECCION DE LOGO (Marca Blanca) ---
    import os
    logo_path = os.path.join('uploads', 'logos', f"{company_id}.png")
    tiene_logo = os.path.exists(logo_path)

    empresa = Company.query.get(company_id)
    return render_template('dashboard.html', 
                           total_equipos=total_equipos,
                           total_registros=total_registros,
                           equipos_asignados=equipos_asignados,
                           equipos_bodega=equipos_bodega,
                           total_completos=total_completos,
                           porcentaje_verificado=porcentaje_verificado,
                           equipos=equipos,
                           registros=registros,
                           proyectos=proyectos,
                           usuarios=usuarios,
                           empresa=empresa,
                           tiempo_ahorrado=tiempo_ahorrado_texto,
                           tiene_logo=tiene_logo)

# --- ENDPOINTS DE MARCA BLANCA ---
from flask import send_from_directory

@app.route('/api/admin/editar_logo', methods=['POST'])
def editar_logo():
    if session.get('usuario_rol') != 'Admin':
        return "Acceso denegado", 403
        
    company_id = session.get('company_id')
    file = request.files.get('logo')
    
    if file and file.filename != '':
        import os
        upload_dir = os.path.join('uploads', 'logos')
        os.makedirs(upload_dir, exist_ok=True)
        # Forzar guardado como empresa_id.png
        filename = f"{company_id}.png"
        file.save(os.path.join(upload_dir, filename))
        flash("Logo corporativo actualizado con éxito.", "success")
    else:
        flash("No se seleccionó ningún archivo válido.", "warning")
        
    return redirect(url_for('admin'))

@app.route('/uploads/logos/<filename>')
def serve_logo(filename):
    import os
    return send_from_directory(os.path.join('uploads', 'logos'), filename)

@app.route('/api/admin/crear_proyecto', methods=['POST'])
def crear_proyecto():
    if session.get('usuario_rol') != 'Admin':
        return "Acceso denegado", 403
    nombre = request.form.get('nombre')
    campos = request.form.get('campos_ocr', 'S/N,MAC')
    if nombre:
        company_id = session.get('company_id')
        nuevo = Proyecto(nombre=nombre, campos_ocr=campos, company_id=company_id)
        db.session.add(nuevo)
        db.session.commit()
    return redirect(url_for('admin') + '#tab-proyectos')

@app.route('/api/admin/editar_proyecto', methods=['POST'])
def editar_proyecto():
    if session.get('usuario_rol') != 'Admin':
        return "Acceso denegado", 403
    
    proyecto_id = request.form.get('proyecto_id')
    nombre = request.form.get('nombre')
    campos = request.form.get('campos_ocr', 'S/N,MAC')
    
    if proyecto_id:
        company_id = session.get('company_id')
        proyecto = Proyecto.query.filter_by(id=proyecto_id, company_id=company_id).first()
        if proyecto:
            if nombre: proyecto.nombre = nombre
            if campos: proyecto.campos_ocr = campos
            
            db.session.commit()
            flash("Proyecto actualizado correctamente", "success")
            
    return redirect(url_for('admin') + '#tab-proyectos')

@app.route('/api/admin/eliminar_proyecto', methods=['POST'])
def eliminar_proyecto():
    if session.get('usuario_rol') != 'Admin':
        return "Acceso denegado", 403
    
    proyecto_id = request.form.get('proyecto_id')
    company_id = session.get('company_id')
    proyecto = Proyecto.query.filter_by(id=proyecto_id, company_id=company_id).first()
    if proyecto:
        db.session.delete(proyecto)
        db.session.commit()
        flash("Proyecto eliminado correctamente", "success")
    return redirect(url_for('admin') + '#tab-proyectos')

@app.route('/api/admin/crear_usuario', methods=['POST'])
def crear_usuario():
    if session.get('usuario_rol') != 'Admin':
        return "Acceso denegado", 403
    correo = request.form.get('correo')
    password = request.form.get('password')
    
    if correo and password:
        company_id = session.get('company_id')
        empresa = Company.query.get(company_id)
        
        # Contar operadores actuales
        conteo_operadores = Usuario.query.filter_by(company_id=company_id, rol='Operador').count()
        
        if empresa and conteo_operadores >= empresa.operadores_limit:
            flash(f"Límite alcanzado: Tu plan permite máximo {empresa.operadores_limit} operadores.", "error")
            return redirect(url_for('admin') + '#tab-usuarios')

        # Check if exists
        exist = Usuario.query.filter_by(correo=correo).first()
        if not exist:
            nuevo = Usuario(correo=correo, password_hash=generate_password_hash(password), rol='Operador', company_id=company_id)
            db.session.add(nuevo)
            db.session.commit()
            flash("Operador creado exitosamente.", "success")
    return redirect(url_for('admin') + '#tab-usuarios')

@app.route('/api/admin/editar_usuario', methods=['POST'])
def editar_usuario():
    if session.get('usuario_rol') != 'Admin':
        return "Acceso denegado", 403
    
    usuario_id = request.form.get('usuario_id')
    correo = request.form.get('correo')
    password = request.form.get('password')
    rol = request.form.get('rol', 'Operador') # Default to Operador if not provided

    if usuario_id:
        usuario = Usuario.query.filter_by(id=usuario_id, company_id=session.get('company_id')).first()
        if usuario:
            if correo:
                # Verificar que el nuevo correo no esté repetido, excluyendo el usuario actual
                exist = Usuario.query.filter(Usuario.correo==correo, Usuario.id!=usuario_id).first()
                if not exist:
                    usuario.correo = correo
            if password:
                usuario.password_hash = generate_password_hash(password)
            
            # Solo permitir cambiar el rol si no es el admin principal editandose a si mismo para quitarse el admin accidentalmente
            if str(usuario_id) != str(session.get('usuario_id')):
                usuario.rol = rol

            db.session.commit()
            flash("Operador actualizado correctamente", "success")
            
    return redirect(url_for('admin') + '#tab-usuarios')

@app.route('/api/admin/eliminar_usuario', methods=['POST'])
def eliminar_usuario():
    if session.get('usuario_rol') != 'Admin':
        return "Acceso denegado", 403
    
    usuario_id = request.form.get('usuario_id')
    # Prevenir que el admin por defecto se borre a sí mismo accidentalmente
    if str(usuario_id) == str(session.get('usuario_id')):
        flash("No puedes eliminar tu propio usuario activo", "error")
        return redirect(url_for('admin') + '#tab-usuarios')
        
    usuario = Usuario.query.filter_by(id=usuario_id, company_id=session.get('company_id')).first()
    if usuario and usuario.rol != 'Admin':
        db.session.delete(usuario)
        db.session.commit()
        flash("Operador eliminado correctamente", "success")
    return redirect(url_for('admin') + '#tab-usuarios')

@app.route('/api/admin/editar_perfil', methods=['POST'])
def editar_perfil():
    if session.get('usuario_rol') != 'Admin':
        return "Acceso denegado", 403
        
    company_id = session.get('company_id')
    usuario_id = session.get('usuario_id')
    
    nuevo_nombre = request.form.get('nombre_empresa', '').strip()
    nuevo_password = request.form.get('password', '').strip()
    
    empresa = Company.query.get(company_id)
    usuario = Usuario.query.get(usuario_id)
    
    if empresa and nuevo_nombre:
        empresa.name = nuevo_nombre
        
    if usuario and nuevo_password:
        usuario.password_hash = generate_password_hash(nuevo_password)
        
    db.session.commit()
    flash("Perfil actualizado correctamente", "success")
    return redirect(url_for('admin'))

@app.route('/api/admin/editar_equipo', methods=['POST'])
def editar_equipo():
    if session.get('usuario_rol') != 'Admin':
        return "Acceso denegado", 403
    
    equipo_id = request.form.get('equipo_id')
    company_id = session.get('company_id')
    equipo = Equipo.query.filter_by(id=equipo_id, company_id=company_id).first()
    if equipo:
        equipo.sn = request.form.get('sn')
        equipo.mac = request.form.get('mac')
        equipo.nombre = request.form.get('nombre')
        equipo.ubicacion = request.form.get('ubicacion')
        equipo.usuario = request.form.get('usuario')
        equipo.estado = request.form.get('estado')
        try:
            db.session.commit()
        except IntegrityError:
            db.session.rollback()
            flash("Error: El número de serie (S/N) ya existe en otro equipo.", "danger")
            return redirect(url_for('admin') + '#tab-inventario')
    return redirect(url_for('admin') + '#tab-inventario')

@app.route('/api/admin/config_ocr', methods=['GET', 'POST'])
def config_ocr():
    if session.get('usuario_rol') != 'Admin':
        return jsonify({"status": "error", "message": "Acceso denegado"}), 403
        
    company_id = session.get('company_id')
    config = OcrConfig.query.filter_by(company_id=company_id).first()
    
    if request.method == 'GET':
        if not config:
            return jsonify({
                "sn_prefix": "",
                "sn_length": "",
                "require_mac": False,
                "require_pn": False,
                "require_ean": False
            })
        return jsonify({
            "sn_prefix": config.sn_prefix or "",
            "sn_length": config.sn_length or "",
            "require_mac": config.require_mac,
            "require_pn": config.require_pn,
            "require_ean": config.require_ean,
            "campos_personalizados": config.campos_personalizados or ""
        })
        
    if request.method == 'POST':
        data = request.json
        if not config:
            config = OcrConfig(company_id=company_id)
            db.session.add(config)
            
        config.sn_prefix = data.get('sn_prefix')
        
        try:
            length_val = data.get('sn_length')
            config.sn_length = int(length_val) if length_val else None
        except ValueError:
            config.sn_length = None
            
        config.require_mac = bool(data.get('require_mac'))
        config.require_pn = bool(data.get('require_pn'))
        config.require_ean = bool(data.get('require_ean'))
        config.campos_personalizados = data.get('campos_personalizados', '').strip() or None
        
        db.session.commit()
        return jsonify({"status": "success", "message": "Reglas OCR actualizadas"})

@app.route('/api/admin/eliminar_equipo', methods=['POST'])
def eliminar_equipo():
    if session.get('usuario_rol') != 'Admin':
        return jsonify({"status": "error", "message": "Acceso denegado"}), 403
    
    equipo_id = request.form.get('equipo_id')
    company_id = session.get('company_id')
    equipo = Equipo.query.filter_by(id=equipo_id, company_id=company_id).first()
    
    if equipo:
        db.session.delete(equipo)
        db.session.commit()
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.args.get('ajax'):
            return jsonify({"status": "success", "message": "Equipo eliminado correctamente"})
            
    return redirect(url_for('admin') + '#tab-inventario')

@app.route('/api/admin/limpiar_no_telefonos', methods=['POST'])
def limpiar_no_telefonos():
    if session.get('usuario_rol') != 'Admin':
        return jsonify({"status": "error", "message": "Acceso denegado"}), 403
        
    company_id = session.get('company_id')
    
    # Eliminar equipos sin MAC de esta empresa
    equipos_a_borrar = Equipo.query.filter_by(company_id=company_id).filter((Equipo.mac == None) | (Equipo.mac == '')).all()
    count_e = len(equipos_a_borrar)
    for e in equipos_a_borrar:
        db.session.delete(e)
        
    # También registros OCR
    registros_a_borrar = RegistroOCR.query.filter_by(company_id=company_id).filter((RegistroOCR.mac == None) | (RegistroOCR.mac == '')).all()
    count_r = len(registros_a_borrar)
    for r in registros_a_borrar:
        db.session.delete(r)
        
    db.session.commit()
    
    return jsonify({
        "status": "success", 
        "message": f"Limpieza completada. Se eliminaron {count_e} equipos y {count_r} registros sin MAC.",
        "count_e": count_e,
        "count_r": count_r
    })

@app.route('/api/admin/exportar_excel')
def exportar_excel():
    if session.get('usuario_rol') != 'Admin':
        return "Acceso denegado", 403
        
    company_id = session.get('company_id')
    equipos = Equipo.query.filter_by(company_id=company_id).order_by(Equipo.fecha_actualizacion.desc()).all()
    
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"
    
    # Cabeceras
    headers = ["S/N", "MAC", "Nombre", "Ubicación", "Usuario", "Estado", "Fecha Actualización"]
    ws.append(headers)
    
    # Estilo Cabeceras (Azul Corporativo #1E3A8A)
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    alignment = Alignment(horizontal="center", vertical="center")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = alignment
        
    # Filas
    for eq in equipos:
        fecha = eq.fecha_actualizacion.strftime('%d/%m/%Y %H:%M') if eq.fecha_actualizacion else '-'
        ws.append([eq.sn, eq.mac or '-', eq.nombre or '-', eq.ubicacion or '-', eq.usuario or '-', eq.estado or '-', fecha])
        
    # Auto-ajustar columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 3

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"Inventario_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# API que recibe SÓLO la imagen, y el servidor lee el Excel local.
@app.route('/api/procesar', methods=['POST'])
def procesar():
    # --- VALIDACIÓN DE SUSCRIPCIÓN ---
    company_id = session.get('company_id')
    empresa = Company.query.get(company_id)
    
    if empresa:
        # 1. Validar expiración de tiempo (Solo Freemium)
        if empresa.plan_type == 'Freemium' and empresa.trial_ends_at and datetime.now() > empresa.trial_ends_at:
            return jsonify({"status": "error", "message": "Tu prueba gratuita de 14 días ha expirado. ¡Actualiza para seguir escaneando!"}), 402
            
        # 2. Validar límite de escaneos (Solo Freemium y Pro)
        if empresa.plan_type != 'Premium':
            total_escaneos = RegistroOCR.query.filter_by(company_id=company_id).count()
            if total_escaneos >= empresa.scans_quota:
                return jsonify({"status": "error", "message": f"Has alcanzado el límite de {empresa.scans_quota} escaneos de tu plan. ¡Actualiza tu suscripción!"}), 402

    if 'foto' not in request.files:
        return jsonify({"status": "error", "message": "Falta la fotografía"}), 400
        
    foto = request.files['foto']
    
    if foto.filename == '':
        return jsonify({"status": "error", "message": "Archivo de foto no seleccionado"}), 400

    # Guardar temporalmente la foto enviada por el celular
    ruta_foto = os.path.join(app.config['UPLOAD_FOLDER'], 'temp_foto.jpg')
    foto.save(ruta_foto)
    
    # --- FIX DE IMÁGENES MÓVILES PARA OPENCV ---
    # Limpiar metadata EXIF de orientación y redimensionar
    try:
        img = Image.open(ruta_foto)
        img = ImageOps.exif_transpose(img) # Corrige rotación si la hay
        
        # Redimensionar si es muy grande (evitar crash en easyocr cv::resize)
        img.thumbnail((1920, 1920), Image.Resampling.LANCZOS)
        
        # Guardarla limpia lista para OCR
        img.convert('RGB').save(ruta_foto, "JPEG", quality=85)
    except Exception as e:
        print(f"Advertencia al pre-procesar imagen: {e}")
        pass
        
    # IMPORTANTE: Definimos la ruta de tu Excel real guardado en tu Compu
    ruta_excel_local = r"c:\Users\usuario\Documents\Proyectos\IN_AN\numeros de serie.xlsx"
    
    # Obtener configuración del proyecto (OCR Flexible)
    proyecto_id = request.form.get('proyecto_id')
    proyecto = Proyecto.query.get(proyecto_id) if proyecto_id else Proyecto.query.first()
    
    # --- REGLAS DINÁMICAS OCR ---
    config_ocr = OcrConfig.query.filter_by(company_id=company_id).first()
    
    # 1. Decidir que campos buscar en la imagen
    if config_ocr:
        campos_list = ["S/N"]
        if config_ocr.require_mac: campos_list.append("MAC")
        if config_ocr.require_pn: campos_list.append("P/N")
        if config_ocr.require_ean: campos_list.append("EAN")
        if config_ocr.campos_personalizados:
             campos_list.extend([c.strip() for c in config_ocr.campos_personalizados.split(",") if c.strip()])
        campos_config = ",".join(campos_list)
    else:
        campos_config = proyecto.campos_ocr if proyecto else "S/N,MAC"

    try:
        # 2. Construir dicccionario de reglas para inyectar al OCR
        reglas_dict = {
            "sn_prefix": config_ocr.sn_prefix if config_ocr else None,
            "sn_length": config_ocr.sn_length if config_ocr else None
        }
        
        # Extraer TODOS los equipos presentes en la foto con OCR flexible + Reglas
        equipos = ocr_script.procesar_lote_cajas(ruta_foto, campos_config=campos_config, reglas=reglas_dict)
        
        # 2. Obtener Workstation ID si viene del mapa
        workstation_id = request.form.get('workstation_id')
        
        if not equipos:
            # Extraer imagen en base64 para previsualización aun si falló el OCR
            img_b64 = None
            try:
                import base64
                with open(ruta_foto, "rb") as image_file:
                    img_b64 = base64.b64encode(image_file.read()).decode('utf-8')
            except:
                pass

            return jsonify({
                "status": "warning", 
                "message": "Fallo de OCR (posible foto múltiple o borrosa). Escribe el S/N y MAC manualmente.",
                "sn": "",
                "mac": "",
                "imagen_b64": img_b64
            })
            
        # 3. Leer metadatos opcionales del formulario
        ubicacion = request.form.get('ubicacion', '').strip() or None
        usuario   = request.form.get('usuario', '').strip() or None
        estado    = request.form.get('estado', '').strip() or None
        forzar    = request.form.get('forzar_actualizacion', 'false').lower() == 'true'
        form_proyecto_id = int(proyecto_id) if proyecto_id and str(proyecto_id).isdigit() else None
        
        # 4. Registrar CADA equipo detectado
        resultados = []
        for eq in equipos:
            sn_telefono  = eq["sn"]
            mac_telefono = eq["mac"]

            resultado = ocr_script.actualizar_inventario_web(
                sn_telefono, mac_telefono, ruta_excel_local,
                ubicacion=ubicacion, usuario=usuario, estado=estado,
                forzar_actualizacion=forzar
            )
            
            # Sincronizar con SQLAlchemy (Supabase)
            resultado.setdefault("sn", sn_telefono)
            resultado.setdefault("mac", mac_telefono)
            
            # Heredar compañía del proyecto o de la sesión activa
            active_company_id = proyecto.company_id if proyecto else session.get('company_id')
            sincronizar_resultado_con_bd(resultado, ubicacion, usuario, estado, form_proyecto_id, company_id=active_company_id)
            
            resultados.append(resultado)
        
        # 5. Respuesta
        img_b64 = None
        try:
            import base64
            with open(ruta_foto, "rb") as image_file:
                img_b64 = base64.b64encode(image_file.read()).decode('utf-8')
        except:
            pass

        if len(resultados) == 1:
            if img_b64: resultados[0]['imagen_b64'] = img_b64
            return jsonify(resultados[0])
        else:
            registrados = sum(1 for r in resultados if r["status"] == "success")
            return jsonify({
                "status": "success",
                "multiples": True,
                "total": len(resultados),
                "registrados": registrados,
                "message": f"Se detectaron {len(resultados)} equipos.",
                "equipos": resultados
            })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": f"Error interno: {str(e)}"}), 500


import base64
import io

# NUEVA API: Procesa un lote (batch) de imágenes en Base64
@app.route('/api/procesar_batch', methods=['POST'])
def procesar_batch():
    import time
    datos = request.json
    if not datos or 'scaneos' not in datos:
        return jsonify({"status": "error", "message": "Falta el array de scaneos"}), 400
        
    scaneos = datos['scaneos']
    if not isinstance(scaneos, list) or len(scaneos) == 0:
        return jsonify({"status": "error", "message": "Array de scaneos vacío"}), 400
        
    ruta_excel_local = r"c:\Users\usuario\Documents\Proyectos\IN_AN\numeros de serie.xlsx"
    batch_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'batch')
    os.makedirs(batch_dir, exist_ok=True)
    
    resultados_totales = []
    
    try:
        for i, scan in enumerate(scaneos):
            try:
                cod_b64 = scan.get('foto_b64', '')
                meta = scan.get('metadata', {})
                
                if "," in cod_b64: cod_b64 = cod_b64.split(",")[1]
                img_data = base64.b64decode(cod_b64)
                img = Image.open(io.BytesIO(img_data))
                img = ImageOps.exif_transpose(img)
                img.thumbnail((1920, 1920), Image.Resampling.LANCZOS)
                
                # Nombre único usando timestamp real de Python
                ts = int(time.time() * 1000)
                ruta_temp = os.path.join(batch_dir, f'img_{i}_{ts}.jpg')
                img.convert('RGB').save(ruta_temp, "JPEG", quality=85)
                
                sn_manual = meta.get('sn', '').strip()
                mac_manual = meta.get('mac', '').strip()
                
                # 1. OCR clásico del equipo
                detectados = ocr_script.procesar_lote_cajas(ruta_temp)
                
                # 2. Si el usuario ingresó manualmente en la Cartilla, lo combinamos/priorizamos
                if sn_manual:
                    if detectados:
                        detectados[0]["sn"] = sn_manual
                        if mac_manual: detectados[0]["mac"] = mac_manual
                    else:
                        detectados = [{"sn": sn_manual, "mac": mac_manual}]
                elif mac_manual:
                    if detectados:
                        detectados[0]["mac"] = mac_manual
                    else:
                        detectados = [{"sn": "DESCONOCIDO", "mac": mac_manual}]
                        
                for eq in detectados:
                    res = ocr_script.actualizar_inventario_web(
                        eq["sn"], eq["mac"], ruta_excel_local,
                        ubicacion=meta.get('ubicacion'), 
                        usuario=meta.get('usuario'), 
                        estado=meta.get('estado'),
                        forzar_actualizacion=True
                    )
                    res.setdefault("sn", eq["sn"])
                    res.setdefault("mac", eq["mac"])
                    
                    # Sincronizar BD
                    proj_id = int(meta.get('proyecto_id')) if meta.get('proyecto_id') and str(meta.get('proyecto_id')).isdigit() else None
                    proyecto_batch = Proyecto.query.get(proj_id) if proj_id else None
                    active_company_id = proyecto_batch.company_id if proyecto_batch else session.get('company_id')
                    
                    # Obtener reglas OCR para este batch
                    config_batch = OcrConfig.query.filter_by(company_id=active_company_id).first()
                    reglas_batch = {
                        "sn_prefix": config_batch.sn_prefix if config_batch else None,
                        "sn_length": config_batch.sn_length if config_batch else None
                    }
                    campos_batch = config_batch.campos_personalizados if (config_batch and config_batch.campos_personalizados) else "S/N,MAC"

                    # 1. Re-procesar con REGLAS si es necesario (o asegurar que se detecten bien)
                    detectados = ocr_script.procesar_lote_cajas(ruta_temp, campos_config=campos_batch, reglas=reglas_batch)

                    for eq in detectados:
                        res = ocr_script.actualizar_inventario_web(
                            eq["sn"], eq["mac"], ruta_excel_local,
                            ubicacion=meta.get('ubicacion'), 
                            usuario=meta.get('usuario'), 
                            estado=meta.get('estado'),
                            forzar_actualizacion=True
                        )
                        res.setdefault("sn", eq["sn"])
                        res.setdefault("mac", eq["mac"])
                        
                        sincronizar_resultado_con_bd(
                            res, 
                            meta.get('ubicacion'), 
                            meta.get('usuario'), 
                            meta.get('estado'), 
                            proj_id,
                            company_id=active_company_id
                        )
                        
                        if res.get('status') == 'success':
                            resultados_totales.append(res)
                    
            except Exception as e:
                print(f"Error procesando scaneo {i}: {e}")

        return jsonify({
            "status": "success",
            "total_procesados": len(resultados_totales),
            "message": f"Sincronizados {len(resultados_totales)} equipos detectados."
        })
    except Exception as e:
        return jsonify({"status": "error", "message": f"Error crítico en lote: {str(e)}"}), 500

# API para actualizar o forzar el registro de un equipo (modificado para SaaS)
@app.route('/api/actualizar_meta', methods=['POST'])
def actualizar_meta():
    """
    Recibe sn, mac, ubicacion, usuario, estado desde el formulario
    de la interfaz (ya sea por duplicado o por S/N no encontrado).
    """
    sn = request.form.get('sn', '').strip()
    mac = request.form.get('mac', '').strip() or None
    ubicacion = request.form.get('ubicacion', '').strip() or None
    usuario   = request.form.get('usuario', '').strip() or None
    estado    = request.form.get('estado', '').strip() or None
    proyecto_id = request.form.get('proyecto_id')
    proyecto_id = int(proyecto_id) if proyecto_id and proyecto_id.isdigit() else None
    proyecto = Proyecto.query.get(proyecto_id) if proyecto_id else None
    active_company_id = proyecto.company_id if proyecto else session.get('company_id')

    if not sn:
        return jsonify({"status": "error", "message": "Falta el número de serie (S/N)."}), 400

    ruta_excel_local = r"c:\Users\usuario\Documents\Proyectos\IN_AN\numeros de serie.xlsx"

    try:
        if mac:
            # Modo Inserción/Corrección Forzada: registra S/N, MAC y Metadata
            resultado = ocr_script.actualizar_inventario_web(sn, mac, ruta_excel_local, 
                                                              ubicacion=ubicacion, 
                                                              usuario=usuario, 
                                                              estado=estado,
                                                              forzar_actualizacion=True)
            # Sincronizamos con PostgreSQL
            resultado.setdefault('sn', sn)
            resultado.setdefault('mac', mac)
            sincronizar_resultado_con_bd(resultado, ubicacion, usuario, estado, proyecto_id, company_id=active_company_id)
        else:
            # Modo Solo Metadatos: solo actualiza ubicacion/usuario/estado
            resultado = ocr_script.actualizar_solo_metadatos(sn, ruta_excel_local, 
                                                              ubicacion=ubicacion, 
                                                              usuario=usuario, 
                                                              estado=estado)
            # Sincronizamos con PostgreSQL
            resultado.setdefault('sn', sn)
            resultado.setdefault('mac', 'ACTUALIZACION_METADATA')
            sincronizar_resultado_con_bd(resultado, ubicacion, usuario, estado, proyecto_id, company_id=active_company_id)
        
        return jsonify(resultado)
    except Exception as e:
        return jsonify({"status": "error", "message": f"Error: {str(e)}"}), 500

# API para registro manual de un equipo desde el Dashboard
@app.route('/api/admin/crear_equipo_manual', methods=['POST'])
def crear_equipo_manual():
    if session.get('usuario_rol') != 'Admin':
        return "Acceso denegado", 403

    sn = request.form.get('sn', '').strip()
    mac = request.form.get('mac', '').strip() or None
    nombre = request.form.get('nombre', '').strip() or 'Equipo Genérico'
    ubicacion = request.form.get('ubicacion', '').strip() or None
    usuario = request.form.get('usuario', '').strip() or None
    estado = request.form.get('estado', 'En Bodega').strip()
    
    if not sn:
        flash("El Número de Serie es obligatorio.", "error")
        return redirect(url_for('admin') + '#tab-inventario')

    try:
        # Guardamos en Database
        company_id = session.get('company_id')
        nuevo_eq = Equipo.query.filter_by(sn=sn, company_id=company_id).first()
        if not nuevo_eq:
            nuevo_eq = Equipo(sn=sn, mac=mac, nombre=nombre, ubicacion=ubicacion, usuario=usuario, estado=estado, company_id=company_id)
            db.session.add(nuevo_eq)
            
            # Registrar auditoría de creación manual
            log = AuditLog(action_type="MANUAL_ENTRY", details={"sn": sn, "mac": mac, "estado": estado}, company_id=company_id)
            db.session.add(log)
        else:
            # Si ya existía, actualizamos datos
            nuevo_eq.mac = mac or nuevo_eq.mac
            nuevo_eq.nombre = nombre or nuevo_eq.nombre
            nuevo_eq.ubicacion = ubicacion or nuevo_eq.ubicacion
            nuevo_eq.usuario = usuario or nuevo_eq.usuario
            nuevo_eq.estado = estado or nuevo_eq.estado
        
        db.session.commit()
        
        # Opcional: Escribir en Excel usando ocr_script.actualizar_inventario_web
        # Aprovechamos la misma función para mantener la data sincronizada
        try:
            ruta_excel_local = app.config.get('EXCEL_PATH', r"c:\Users\usuario\Documents\Proyectos\IN_AN\numeros de serie.xlsx")
            ocr_script.actualizar_inventario_web(
                sn, mac, ruta_excel_local,
                ubicacion=ubicacion, usuario=usuario, estado=estado, forzar_actualizacion=True
            )
        except Exception as ex_excel:
            print(f"Alerta: no se pudo escribir en el excel: {ex_excel}")

        flash(f"Equipo {sn} registrado correctamente de forma manual.", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Error al registrar equipo manual: {str(e)}", "error")

    return redirect(url_for('admin') + '#tab-inventario')

# API para descargar el Excel ya modificado
@app.route('/api/descargar_excel', methods=['GET'])
def descargar_excel():
    ruta_excel = os.path.join(app.config['UPLOAD_FOLDER'], 'temp_excel.xlsx')
    if os.path.exists(ruta_excel):
        return send_file(ruta_excel, as_attachment=True, download_name="Inventario_Actualizado.xlsx")
    return "Archivo no encontrado", 404
# --- INTEGRACIÓN MERCADOPAGO (CHILE) ---
MP_ACCESS_TOKEN = os.environ.get('MP_ACCESS_TOKEN', 'TEST-252174291722881-032009-4ce4356499691b0dc37269b51ebd29c7-147372202') # Reemplazar por token real en prod

@app.route('/api/admin/crear_preferencia', methods=['POST'])
def crear_preferencia():
    if 'usuario_id' not in session or session.get('usuario_rol') != 'Admin':
        return jsonify({"status": "error", "message": "Acceso denegado"}), 403
        
    data = request.json
    plan = data.get('plan_type', 'pro').lower()
    
    # Precios nominales en CLP (Ejemplo)
    planes = {
        'pro': {"title": "Jet OCR - Plan PRO", "price": 18000},    # ~$19.99 USD
        'premium': {"title": "Jet OCR - Plan PREMIUM", "price": 36000} # ~$39.99 USD
    }
    
    info_plan = planes.get(plan, planes['pro'])
    
    headers = {
        "Authorization": f"Bearer {MP_ACCESS_TOKEN}",
        "Content-Type": "application/json"
    }
    
    # URLs de retorno (Ajustar cuando tengas dominio)
    host_url = request.host_url.rstrip('/')
    company_id_str = str(session.get('company_id'))
    
    preference_data = {
        "items": [
            {
                "title": info_plan["title"],
                "quantity": 1,
                "currency_id": "CLP",
                "unit_price": info_plan["price"]
            }
        ],
        "back_urls": {
            "success": f"{host_url}/api/pago_exitoso?plan={plan}&company_id={company_id_str}",
            "failure": f"{host_url}/admin#plansModal",
            "pending": f"{host_url}/admin#plansModal"
        },
        "auto_return": "approved",
        "external_reference": company_id_str
    }
    
    try:
        response = requests.post(
            "https://api.mercadopago.com/checkout/preferences",
            headers=headers,
            json=preference_data
        )
        res_data = response.json()
        if response.status_code == 201 or response.status_code == 200:
            return jsonify({"status": "success", "url": res_data.get("init_point")})
        else:
            print(f"Error MercadoPago: {res_data}")
            return jsonify({"status": "error", "message": "No se pudo generar el enlace de pago."}), 500
    except Exception as e:
         return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/pago_exitoso', methods=['GET'])
def pago_exitoso():
    plan = request.args.get('plan', 'pro').lower()
    company_id = request.args.get('company_id') or session.get('company_id')
    
    if not company_id:
        flash("Error: No se pudo identificar la empresa asociada al pago.", "danger")
        return redirect(url_for('admin'))
        
    empresa = Company.query.get(company_id)
    if empresa:
        if plan == 'pro':
            empresa.plan_type = 'Pro'
            empresa.scans_quota = 1000
        elif plan == 'premium':
            empresa.plan_type = 'Premium'
            empresa.scans_quota = 999999
            
        db.session.commit()
        flash("¡Felicidades! Tu Plan ha sido actualizado correctamente.", "success")
    else:
        flash("Error al actualizar plan en la base de datos.", "danger")
        
    return redirect(url_for('admin'))


if __name__ == '__main__':
    with app.app_context():
        inicializar_db()
    # host='0.0.0.0' permite que otros dispositivos en la red local (el celular) accedan a la página
    is_debug = os.environ.get('FLASK_DEBUG', 'True').lower() == 'true'
    app.run(host='0.0.0.0', port=5000, debug=is_debug)
