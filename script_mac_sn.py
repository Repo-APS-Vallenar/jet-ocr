import easyocr
import re
import math
import openpyxl
import os

# Inicializar el lector (se carga una sola vez)
# Usamos verbose=False para evitar problemas de codificación Unicode en consola Windows al descargar modelos
reader = easyocr.Reader(['es', 'en'], verbose=False)

def calcular_centro(bbox):
    """Calcula el punto central (x, y) de las coordenadas devueltas por EasyOCR."""
    x_coords = [p[0] for p in bbox]
    y_coords = [p[1] for p in bbox]
    return (sum(x_coords)/4, sum(y_coords)/4)

def calcular_distancia(centro1, centro2):
    """Calcula la distancia espacial entre dos textos en la imagen."""
    return math.sqrt((centro1[0] - centro2[0])**2 + (centro1[1] - centro2[1])**2)

def procesar_lote_cajas(ruta_imagen, campos_config="S/N,MAC", reglas=None):
    """
    Lee una foto con múltiples equipos y extrae los campos definidos en campos_config.
    campos_config: string ej 'S/N,MAC' o 'IMEI,Activo'
    reglas: dict con prefijos o longitudes específicas para validar S/N
    """
    print(f"Procesando imagen con configuración OCR [{campos_config}]: {ruta_imagen}...")
    results = reader.readtext(ruta_imagen)
    
    config = [c.strip().upper() for c in campos_config.split(',')]
    
    # Patrones por defecto (Yealink / Estándar)
    # Palabras prohibidas (Bloqueo total para evitar 'COMPATIBLEWITH...' y similares)
    BLOCKLIST = ["COMPATIBLE", "MADEIN", "PRODUCT", "MODEL", "FCCID", "ICID", "ASSEMBL", "CHINA", "VIETNAM", "CONTAINS", "MODULO", "MODULE"]
    
    # Listas temporales de hallazgos con prioridad
    hallazgos_sn = []
    hallazgos_mac = []
    
    # 1. Encontrar todos los textos y categorizarlos por "calidad"
    for bbox, text, prob in results:
        t_clean = text.replace(" ", "").upper()
        # Si contiene palabras prohibidas, ignoramos este bloque de texto por completo
        if any(word in t_clean for word in BLOCKLIST):
            continue
            
        # --- DETECCIÓN DE MAC CON FORMATO (Colons o Dashes) ---
        if "MAC" in config:
            match_fmt = re.match(r'^([0-9A-Z]{2}[:\-]){5}[0-9A-Z]{2}$', t_clean)
            if match_fmt:
                val_mac = re.sub(r"[^A-Z0-9]", "", t_clean).replace('O', '0')
                if len(val_mac) == 12:
                    hallazgos_mac.append({"valor": val_mac, "centro": calcular_centro(bbox), "prioridad": 10})
                    continue

        t_clean_raw = re.sub(r"[^A-Z0-9]", "", t_clean) # Versión hiper-limpia
        if len(t_clean_raw) < 4: continue

        # --- DETECCIÓN DE MAC ---
        if "MAC" in config:
            # Prefijo explícito MAC (Tolerante a Alphanum ej. O->0)
            match_explicit_mac = re.search(r'^MAC([A-Z0-9]{11,12})$', t_clean_raw)
            if match_explicit_mac:
                val_mac = match_explicit_mac.group(1).replace('O', '0')
                hallazgos_mac.append({"valor": val_mac, "centro": calcular_centro(bbox), "prioridad": 5})
                continue # Ya es MAC confirmada, ignorar resto
                
            # Si no empieza con MAC, pero tiene exactamente 12 HEX
            if len(t_clean_raw) == 12 and re.match(r'^[0-9A-F]{12}$', t_clean_raw):
                hallazgos_mac.append({"valor": t_clean_raw, "centro": calcular_centro(bbox)})
                continue

        # --- DETECCIÓN DE S/N ---
        if "S/N" in config or "SN" in config:
             # A. REGLAS DINÁMICAS (Prioridad Máxima del Cliente)
             if reglas and (reglas.get('sn_prefix') or reglas.get('sn_length')):
                 prefix = reglas.get('sn_prefix')
                 length = reglas.get('sn_length')
                 
                 prefix_matchea = True
                 length_matchea = True
                 
                 val_sn = t_clean_raw.upper().replace('O', '0')
                 
                 if prefix:
                     prefix_tol = prefix.upper().replace('O', '0')
                     prefix_matchea = val_sn.startswith(prefix_tol)
                 if length:
                     length_matchea = len(val_sn) == int(length)
                     
                 if prefix_matchea and length_matchea and len(val_sn) >= 4:
                     hallazgos_sn.append({"valor": val_sn, "centro": calcular_centro(bbox), "prioridad": 10})
                     continue

             # B. PATRONES ESTÁNDAR (Por Defecto / Respaldos)
             # Prioridad 1: Patrón exacto solicitado 30104 (Yealink)
             match_sn_prefijo = re.search(r'([3][0O]1[0O]4[A-Z0-9]+)', t_clean_raw, re.IGNORECASE)
             if match_sn_prefijo:
                 val_sn = match_sn_prefijo.group(1).upper().replace('O', '0')
                 if val_sn.startswith('30104'):
                    hallazgos_sn.append({"valor": val_sn, "centro": calcular_centro(bbox), "prioridad": 8})
                    continue

             # Prioridad 1.5: Patrones Fortinet (FGT, FON, FSW o 16 caracteres alfanuméricos)
             match_fortinet = re.search(r'((FGT|FON|FSW)[A-Z0-9]{13})', t_clean_raw, re.IGNORECASE)
             if match_fortinet:
                 val_sn = match_fortinet.group(1).upper().replace('O', '0')
                 hallazgos_sn.append({"valor": val_sn, "centro": calcular_centro(bbox), "prioridad": 9})
                 continue

             # Prioridad 2: Buscar palabra clave "S/N" o "SN" en el texto
             sn_found_by_kw = False
             for kw in ["S/N", "SN", "SERIE", "SERIAL"]:
                 if kw in t_clean:
                     val_post = t_clean.split(kw)[-1].strip(":- ")
                     val_post_raw = re.sub(r"[^A-Z0-9]", "", val_post)
                     if len(val_post_raw) >= 10:
                        hallazgos_sn.append({"valor": val_post_raw, "centro": calcular_centro(bbox), "prioridad": 5})
                        sn_found_by_kw = True
                        break
             if sn_found_by_kw:
                 continue

             # Prioridad 3: Regex genérico (Detección por longitud Y contener números)
             match_gen = re.match(r'^(?=.*[0-9])[A-Z0-9]{10,21}$', t_clean_raw)
             if match_gen:
                 hallazgos_sn.append({"valor": t_clean_raw, "centro": calcular_centro(bbox), "prioridad": 1})

    # 2. PROCESO DE EMPAREJAMIENTO
    # Ordenar SN por prioridad (los 301044H primero)
    hallazgos_sn.sort(key=lambda x: x["prioridad"], reverse=True)
    
    equipos = []
    
    # Si detectamos MACs, intentamos emparejar
    if "MAC" in config and hallazgos_mac:
        for m in hallazgos_mac:
            mejor_s = None
            menor_distancia = float('inf')
            # Solo buscamos entre los SN disponibles
            for s in hallazgos_sn:
                dist = calcular_distancia(m["centro"], s["centro"])
                if dist < menor_distancia:
                    menor_distancia = dist
                    mejor_s = s
            
            if mejor_s:
                equipos.append({"sn": mejor_s["valor"], "mac": m["valor"]})
                hallazgos_sn.remove(mejor_s)
    
    # Si sobran SNs (especialmente los de prioridad 10), los agregamos aunque no tengan MAC emparejada todavía
    for s in hallazgos_sn:
        # Solo agregamos sobrantes si tienen prioridad alta o si no se detectó ninguna MAC en toda la foto
        if s["prioridad"] >= 5 or not teams_have_mac(equipos):
            # Evitar duplicados de SN en la lista de equipos final
            if not any(e["sn"] == s["valor"] for e in equipos):
                equipos.append({"sn": s["valor"], "mac": ""})
            
    return equipos

def teams_have_mac(teams):
    return any(t["mac"] != "" for t in teams)

def buscar_columna_por_texto(sheet, posibles_textos):
    """Busca en las primeras 5 filas la columna que coincida con alguno de los textos esperados y devuelve su índice."""
    for row in range(1, 6):
        for col in range(1, sheet.max_column + 1):
            val = sheet.cell(row=row, column=col).value
            if val and isinstance(val, str):
                val_clean = val.strip().lower()
                for t in posibles_textos:
                    if t in val_clean:
                        return col, row
    return None, None

NOMBRE_HOJA_LOG = "Registro OCR"

def registrar_en_log(wb, sn, mac, estado_inventario, equipo_nombre=None, ubicacion=None, usuario=None, estado=None):
    """
    Registra UN escaneo en la hoja 'Registro OCR' del workbook ya abierto.
    
    - Crea la hoja con cabeceras si no existe.
    - Si la combinación S/N + MAC ya está en el log, NO la vuelve a insertar.
    - El workbook NO se guarda aquí; el llamador debe hacer wb.save() después.
    
    Retorna True si se insertó, False si ya existía (duplicado en log).
    """
    import datetime

    # Crear hoja si no existe
    if NOMBRE_HOJA_LOG not in wb.sheetnames:
        log_sheet = wb.create_sheet(NOMBRE_HOJA_LOG)
        cabeceras = ["Fecha/Hora", "S/N", "MAC", "Equipo", "En Inventario", "Ubicación", "Usuario", "Estado"]
        for i, cab in enumerate(cabeceras, start=1):
            log_sheet.cell(row=1, column=i).value = cab
    else:
        log_sheet = wb[NOMBRE_HOJA_LOG]

    # Verificar si esta combinación S/N + MAC ya fue registrada (anti-duplicación)
    sn_upper  = str(sn).strip().upper()
    mac_upper = str(mac).strip().upper()

    for row in range(2, log_sheet.max_row + 1):
        sn_celda  = log_sheet.cell(row=row, column=2).value
        mac_celda = log_sheet.cell(row=row, column=3).value
        if sn_celda and mac_celda:
            if str(sn_celda).strip().upper() == sn_upper and \
               str(mac_celda).strip().upper() == mac_upper:
                return False  # Ya existe, no insertar de nuevo

    # Insertar nuevo registro
    nueva_fila = log_sheet.max_row + 1
    log_sheet.cell(row=nueva_fila, column=1).value = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log_sheet.cell(row=nueva_fila, column=2).value = sn
    log_sheet.cell(row=nueva_fila, column=3).value = mac
    log_sheet.cell(row=nueva_fila, column=4).value = equipo_nombre or "-"
    log_sheet.cell(row=nueva_fila, column=5).value = estado_inventario  # "Encontrado", "No encontrado", "No Yealink"
    log_sheet.cell(row=nueva_fila, column=6).value = ubicacion or ""
    log_sheet.cell(row=nueva_fila, column=7).value = usuario or ""
    log_sheet.cell(row=nueva_fila, column=8).value = estado or ""
    return True  # Insertado correctamente

def actualizar_inventario(ruta_imagen, archivo_excel):
    """
    Procesa una foto de un equipo, busca su S/N en el Excel usando openpyxl y actualiza su MAC.
    Mantiene el formato original del archivo excel.
    """
    try:
        if not os.path.exists(archivo_excel):
             print(f"Error: No se encontró el archivo '{archivo_excel}'.")
             return
             
        # 1. El script extrae los datos de la foto
        equipo_encontrado = procesar_lote_cajas(ruta_imagen)
        
        if not equipo_encontrado:
            print("No se detectó claramente la MAC o el S/N en la foto, por favor toma otra imagen más nítida.")
            return

        datos_equipo = equipo_encontrado[0]  # Tomamos el primero detectado en la foto
        sn_telefono = datos_equipo["sn"]
        mac_telefono = datos_equipo["mac"]
        
        print(f"Detectado en foto -> S/N: {sn_telefono} | MAC: {mac_telefono}")

        # 2. Cargar Excel con openpyxl para no perder el estilo
        wb = openpyxl.load_workbook(archivo_excel)
        sheet = wb.active
        
        # Buscar en qué columna y fila está el encabezado del S/N y la MAC
        col_sn, fila_header = buscar_columna_por_texto(sheet, ["n/s", "serie", "sn"])
        if not col_sn:
            print("Error: No se encontró la columna de Números de Serie ('n/s', 'serie', 'sn') en el Excel.")
            return

        # Buscar si ya existe una columna 'MAC'. Si no, la creamos en la siguiente disponible.
        col_mac, _ = buscar_columna_por_texto(sheet, ["mac"])
        if not col_mac:
            col_mac = sheet.max_column + 1
            sheet.cell(row=fila_header, column=col_mac).value = "MAC"

        col_modelo, _ = buscar_columna_por_texto(sheet, ["modelo", "dispositivo", "nombre"])

        # 3. Buscar el S/N escaneado en todas las filas debajo del encabezado (filtrando por YEALINK)
        fila_encontrada = None
        for row in range(fila_header + 1, sheet.max_row + 1):
            celda_sn = sheet.cell(row=row, column=col_sn).value
            if celda_sn and str(celda_sn).strip().upper() == sn_telefono.upper():
                
                # Comprobar que sea un equipo YEALINK verificando la columna modelo
                es_yealink = False
                if col_modelo:
                    val_modelo = sheet.cell(row=row, column=col_modelo).value
                    if val_modelo and "YEALINK" in str(val_modelo).upper():
                        es_yealink = True
                else:
                    # Si no hay columna modelo en el excel, asumimos que puede serlo, 
                    # pero lo ideal es que exista para filtrar correctamente
                    es_yealink = True
                    
                if es_yealink:
                    fila_encontrada = row
                    break
                else:
                    print(f"Info: Se encontró el S/N {sn_telefono}, pero el equipo no es de la marca YEALINK. Se ignora.")

        # 4. Actualizar la MAC si encontramos la fila
        if fila_encontrada:
            sheet.cell(row=fila_encontrada, column=col_mac).value = mac_telefono
            
            # Guardamos
            wb.save(archivo_excel)
            
            # Intentar leer el nombre si existe la columna
            nombre_equipo = "Desconocido"
            if col_modelo:
                val_nombre = sheet.cell(row=fila_encontrada, column=col_modelo).value
                if val_nombre: nombre_equipo = str(val_nombre).strip()
                
            print(f"¡Éxito! Equipo '{nombre_equipo}' (S/N {sn_telefono}) actualizado con la MAC de la foto.")
        else:
             print(f"Advertencia: El S/N {sn_telefono} de la foto NO se encuentra en '{archivo_excel}'.")

    except Exception as e:
         import traceback
         print(f"Error al procesar el Excel: {e}")
         traceback.print_exc()

def actualizar_inventario_web(sn_telefono, mac_telefono, archivo_excel, ubicacion=None, usuario=None, estado=None, forzar_actualizacion=False):
    """
    Función optimizada para el servidor web. Soporta metadatos adicionales.
    Si el S/N ya tiene MAC registrada, devuelve 'duplicate' con los datos existentes.
    Con forzar_actualizacion=True, permite sobreescribir los datos existentes.
    """
    try:
        if not os.path.exists(archivo_excel):
             return {"status": "error", "message": f"No se encontró el archivo '{archivo_excel}' en servidor."}

        # --- FILTRO ANTIDUPLICADOS / CORRECCIÓN OCR ---
        import re
        sn_telefono = str(sn_telefono).strip().upper() if sn_telefono else ""
        mac_telefono = str(mac_telefono).strip().upper() if mac_telefono else ""

        # Normalización temporal para evaluar formato
        cand_sn = sn_telefono.replace("O", "0").replace("I", "1").replace(" ", "").replace(":", "").replace("-", "")
        if cand_sn.startswith("MAC"):
            cand_sn = cand_sn.replace("MAC", "", 1) # Quitar solo el primer prefijo

        # Verificar si cumple patrón exacto de MAC (12 caracteres Hexadecimales)
        es_mac_address = re.match(r'^[0-9A-F]{12}$', cand_sn)

        if es_mac_address:
            # Es una MAC perdida en el campo S/N
            if not mac_telefono or mac_telefono in ["DESCONOCIDO", "FALTA MAC"]:
                mac_telefono = cand_sn
                sn_telefono = "DESCONOCIDO"

        if mac_telefono:
            mac_telefono = mac_telefono.replace("O", "0").replace("I", "1").replace(":", "").replace("-", "").strip()

        wb = openpyxl.load_workbook(archivo_excel)
        sheet = wb.active
        
        col_sn, fila_header = buscar_columna_por_texto(sheet, ["n/s", "serie", "sn"])
        if not col_sn:
            return {"status": "error", "message": "No se encontró la columna de Números de Serie en el Excel."}

        # Obtener o crear cada columna adicional
        def obtener_o_crear_col(nombres_busqueda, nombre_nuevo):
            col, _ = buscar_columna_por_texto(sheet, nombres_busqueda)
            if not col:
                col = sheet.max_column + 1
                sheet.cell(row=fila_header, column=col).value = nombre_nuevo
            return col

        col_mac = obtener_o_crear_col(["mac"], "MAC")
        col_modelo, _ = buscar_columna_por_texto(sheet, ["modelo", "dispositivo", "nombre"])
        col_ubicacion = obtener_o_crear_col(["ubicación", "ubicacion", "box", "oficina"], "Ubicación")
        col_usuario  = obtener_o_crear_col(["usuario", "asignado", "responsable", "user"], "Usuario")
        col_estado   = obtener_o_crear_col(["estado", "status"], "Estado")

        fila_encontrada = None
        es_yealink = False
        nombre_equipo = "Desconocido"
        
        fila_mac_encontrada = None
        sn_de_esa_mac = None
        
        for row in range(fila_header + 1, sheet.max_row + 1):
            celda_sn = sheet.cell(row=row, column=col_sn).value
            celda_mac = sheet.cell(row=row, column=col_mac).value if col_mac else None
            
            if celda_sn and str(celda_sn).strip().upper() == sn_telefono.upper():
                if col_modelo:
                    val_modelo = sheet.cell(row=row, column=col_modelo).value
                    if val_modelo:
                        nombre_equipo = str(val_modelo).strip()
                        if "YEALINK" in nombre_equipo.upper():
                            es_yealink = True
                else:
                    es_yealink = True
                    
                if es_yealink:
                    fila_encontrada = row
                    break
            # Si no coincide el S/N, comprobamos si la MAC coincide (prevenir duplicados por error de OCR)
            elif celda_mac and mac_telefono and str(celda_mac).strip().upper() == mac_telefono.upper():
                fila_mac_encontrada = row
                sn_de_esa_mac = str(celda_sn).strip() if celda_sn else None
                if col_modelo:
                    val_modelo = sheet.cell(row=row, column=col_modelo).value
                    if val_modelo:
                        nombre_equipo = str(val_modelo).strip()
                        
        if fila_encontrada and not es_yealink:
                    # No es Yealink: registrar en Log y retornar
                    registrar_en_log(wb, sn_telefono, mac_telefono, 
                                     "No Yealink", equipo_nombre=nombre_equipo,
                                     ubicacion=ubicacion, usuario=usuario, estado=estado)
                    wb.save(archivo_excel)
                    return {
                        "status": "warning", 
                        "message": f"El S/N coincide, pero '{nombre_equipo}' no es YEALINK. Guardado en hoja '{NOMBRE_HOJA_LOG}'.",
                        "equipo_nombre": nombre_equipo
                    }

        if not fila_encontrada:
            if fila_mac_encontrada:
                # S/N no encontrado, pero la MAC SI. ¡Probable error de OCR en el S/N!
                if forzar_actualizacion:
                    fila_encontrada = fila_mac_encontrada
                    sheet.cell(row=fila_encontrada, column=col_sn).value = sn_telefono
                else:
                    # Devolver como 'duplicate' para que el usuario pueda corregir el S/N si OCR falló
                    mac_existente      = sheet.cell(row=fila_mac_encontrada, column=col_mac).value
                    ubicacion_existente = sheet.cell(row=fila_mac_encontrada, column=col_ubicacion).value
                    usuario_existente   = sheet.cell(row=fila_mac_encontrada, column=col_usuario).value
                    estado_existente    = sheet.cell(row=fila_mac_encontrada, column=col_estado).value
                    
                    registrar_en_log(wb, sn_telefono, mac_telefono, 
                                     "Atención: MAC ya existe bajo otro S/N", equipo_nombre=nombre_equipo,
                                     ubicacion=ubicacion, usuario=usuario, estado=estado)
                    wb.save(archivo_excel)
                    
                    return {
                        "status": "warning", # warning en vez de duplicate para titulo rojo
                        "message": f"Atención: La MAC '{mac_telefono}' ya está registrada, pero con el S/N '{sn_de_esa_mac}'. El OCR parece haberse equivocado leyendo el SN. Corrígelo.",
                        "sn": sn_de_esa_mac, # Mostramos el original guardado para que lo edite si quiere
                        "mac": mac_telefono,
                        "mac_coincide": True,
                        "equipo_nombre": nombre_equipo,
                        "existente": {
                            "mac":       str(mac_existente).strip() if mac_existente else None,
                            "ubicacion": str(ubicacion_existente).strip() if ubicacion_existente else None,
                            "usuario":   str(usuario_existente).strip() if usuario_existente else None,
                            "estado":    str(estado_existente).strip() if estado_existente else None,
                        }
                    }
            else:
                if forzar_actualizacion:
                    # Si forzamos actualización y la MAC tampoco existe, agregamos la fila al final del Excel local
                    fila_encontrada = sheet.max_row + 1
                    sheet.cell(row=fila_encontrada, column=col_sn).value = sn_telefono
                    nombre_equipo = "Agregado Manual (SaaS)"
                else:
                    # Aunque no esté en el inventario, lo registramos en el Log
                    registrar_en_log(wb, sn_telefono, mac_telefono, 
                                     "No encontrado en inventario",
                                     ubicacion=ubicacion, usuario=usuario, estado=estado)
                    wb.save(archivo_excel)
                    return {
                        "status": "not_found", 
                        "message": f"El S/N no pertenece al inventario original. Puedes agregarlo a la fuerza o corregirlo.",
                        "equipo_nombre": "No encontrado",
                        "sn": sn_telefono,
                        "mac": mac_telefono
                    }

        # Leer datos ya existentes en el Excel para esa fila
        mac_existente      = sheet.cell(row=fila_encontrada, column=col_mac).value
        ubicacion_existente = sheet.cell(row=fila_encontrada, column=col_ubicacion).value
        usuario_existente   = sheet.cell(row=fila_encontrada, column=col_usuario).value
        estado_existente    = sheet.cell(row=fila_encontrada, column=col_estado).value

        # Detectar si la fila ya tiene una MAC cargada (duplicado)
        if mac_existente and not forzar_actualizacion:
            # Registrar en Log aunque sea duplicado en inventario
            registrar_en_log(wb, sn_telefono, mac_telefono, 
                             "Duplicado (ya registrado)", equipo_nombre=nombre_equipo,
                             ubicacion=ubicacion, usuario=usuario, estado=estado)
            wb.save(archivo_excel)
            mac_coincide = str(mac_existente).strip().upper() == mac_telefono.upper()
            return {
                "status": "duplicate",
                "message": "Este equipo ya fue registrado anteriormente. Puedes actualizar los campos que faltan.",
                "sn": sn_telefono,
                "mac": mac_telefono,
                "mac_coincide": mac_coincide,
                "equipo_nombre": nombre_equipo,
                "existente": {
                    "mac":       str(mac_existente).strip() if mac_existente else None,
                    "ubicacion": str(ubicacion_existente).strip() if ubicacion_existente else None,
                    "usuario":   str(usuario_existente).strip() if usuario_existente else None,
                    "estado":    str(estado_existente).strip() if estado_existente else None,
                }
            }

        # Registrar nuevo dato (o actualizar forzado)
        # La MAC siempre se graba
        sheet.cell(row=fila_encontrada, column=col_mac).value = mac_telefono
        
        # Los metadatos opcionales: solo se graban si el usuario los envió
        # En actualización parcial, no se pisan los datos que ya existen si no se enviaron nuevos
        if ubicacion:
            sheet.cell(row=fila_encontrada, column=col_ubicacion).value = ubicacion
        if usuario:
            sheet.cell(row=fila_encontrada, column=col_usuario).value = usuario
        if estado:
            sheet.cell(row=fila_encontrada, column=col_estado).value = estado

        # Registrar en Log + guardar en hoja principal
        registrar_en_log(wb, sn_telefono, mac_telefono, 
                         "Registrado", equipo_nombre=nombre_equipo,
                         ubicacion=ubicacion, usuario=usuario, estado=estado)
        wb.save(archivo_excel)
        return {
            "status": "success", 
            "message": "Registrado exitosamente en el inventario.",
            "equipo_nombre": nombre_equipo,
            "sn": sn_telefono,
            "mac": mac_telefono,
        }

    except Exception as e:
         return {"status": "error", "message": str(e)}

def actualizar_solo_metadatos(sn_telefono, archivo_excel, ubicacion=None, usuario=None, estado=None):
    """
    Actualiza únicamente los metadatos (ubicación, usuario, estado) de un equipo ya registrado.
    No requiere procesar ninguna imagen. Ideal para completar datos faltantes.
    """
    try:
        if not os.path.exists(archivo_excel):
            return {"status": "error", "message": f"No se encontró el archivo '{archivo_excel}'."}

        wb = openpyxl.load_workbook(archivo_excel)
        sheet = wb.active
        
        col_sn, fila_header = buscar_columna_por_texto(sheet, ["n/s", "serie", "sn"])
        if not col_sn:
            return {"status": "error", "message": "No se encontró la columna de Números de Serie."}

        def obtener_o_crear_col(nombres_busqueda, nombre_nuevo):
            col, _ = buscar_columna_por_texto(sheet, nombres_busqueda)
            if not col:
                col = sheet.max_column + 1
                sheet.cell(row=fila_header, column=col).value = nombre_nuevo
            return col

        col_ubicacion = obtener_o_crear_col(["ubicación", "ubicacion", "box", "oficina"], "Ubicación")
        col_usuario   = obtener_o_crear_col(["usuario", "asignado", "responsable", "user"], "Usuario")
        col_estado    = obtener_o_crear_col(["estado", "status"], "Estado")

        fila_encontrada = None
        for row in range(fila_header + 1, sheet.max_row + 1):
            celda_sn = sheet.cell(row=row, column=col_sn).value
            if celda_sn and str(celda_sn).strip().upper() == sn_telefono.upper():
                fila_encontrada = row
                break

        if not fila_encontrada:
            return {"status": "error", "message": f"El S/N {sn_telefono} no existe en la base de datos."}

        if ubicacion:
            sheet.cell(row=fila_encontrada, column=col_ubicacion).value = ubicacion
        if usuario:
            sheet.cell(row=fila_encontrada, column=col_usuario).value = usuario
        if estado:
            sheet.cell(row=fila_encontrada, column=col_estado).value = estado

        wb.save(archivo_excel)
        return {
            "status": "success",
            "message": "Datos actualizados correctamente en el inventario.",
            "sn": sn_telefono,
        }

    except Exception as e:
        return {"status": "error", "message": str(e)}

def actualizar_inventario_batch(lista_datos, archivo_excel):
    """
    Función optimizada para el procesamiento OFFLINE por lotes.
    Recibe un arreglo de diccionarios [{'sn': 'xxx', 'mac': 'yyy'}, ...] 
    Busca todas las filas coincidentes y guarda el Excel una sola vez.
    Retorna estadisticas de cuantos se guardaron con exito.
    """
    try:
        if not os.path.exists(archivo_excel):
             return {"status": "error", "message": f"No se encontró el archivo '{archivo_excel}' en servidor."}

        wb = openpyxl.load_workbook(archivo_excel)
        sheet = wb.active
        
        col_sn, fila_header = buscar_columna_por_texto(sheet, ["n/s", "serie", "sn"])
        if not col_sn:
            return {"status": "error", "message": "No se encontró la columna de Números de Serie en el Excel."}

        col_mac, _ = buscar_columna_por_texto(sheet, ["mac"])
        if not col_mac:
            col_mac = sheet.max_column + 1
            sheet.cell(row=fila_header, column=col_mac).value = "MAC"

        col_modelo, _ = buscar_columna_por_texto(sheet, ["modelo", "dispositivo", "nombre"])

        actualizados = 0
        errores = []

        # Convertir lista a diccionario temporal para búsqueda O(1)
        mapa_novedades = {str(item["sn"]).strip().upper(): item["mac"] for item in lista_datos}
        
        for row in range(fila_header + 1, sheet.max_row + 1):
            celda_sn = sheet.cell(row=row, column=col_sn).value
            if celda_sn:
                sn_actual = str(celda_sn).strip().upper()
                
                # Si el SN del Excel está en la tanda de fotos que trajo el celular
                if sn_actual in mapa_novedades:
                    mac_telefono = mapa_novedades[sn_actual]
                    es_yealink = False
                    nombre_equipo = "Desconocido"
                    
                    if col_modelo:
                        val_modelo = sheet.cell(row=row, column=col_modelo).value
                        if val_modelo:
                            nombre_equipo = str(val_modelo).strip()
                            if "YEALINK" in nombre_equipo.upper():
                                es_yealink = True
                    else:
                        es_yealink = True
                        
                    if es_yealink:
                        sheet.cell(row=row, column=col_mac).value = mac_telefono
                        actualizados += 1
                        # Lo sacamos para saber cuáles no se encontraron
                        del mapa_novedades[sn_actual] 
                    else:
                        errores.append(f"El S/N {sn_actual} es un {nombre_equipo} (NO Yealink).")

        # Guardamos el Excel 1 sola vez por todas las 66 fotos
        if actualizados > 0:
            wb.save(archivo_excel)
            
        no_encontrados = list(mapa_novedades.keys())
        
        mensaje = f"Se actualizaron {actualizados} equipos exitosamente."
        if no_encontrados:
            mensaje += f" S/N reportados que no existen en base: {len(no_encontrados)}."
        if errores:
            mensaje += f" Ignorados por no ser Yealink: {len(errores)}."
            
        return {
            "status": "success" if actualizados > 0 else "warning", 
            "message": mensaje,
            "actualizados": actualizados,
            "no_encontrados": len(no_encontrados),
            "ignorados": len(errores)
        }

    except Exception as e:
         return {"status": "error", "message": f"Fallo al guardar Lote: {str(e)}"}

# --- EJEMPLO DE USO ---
if __name__ == "__main__":
    
    # Cambia los nombres por los de tu prueba real
    ARCHIVO_INVENTARIO = "numeros de serie.xlsx" 
    FOTO_EQUIPO = "Serie equipos.jpg"       
    
    actualizar_inventario(FOTO_EQUIPO, ARCHIVO_INVENTARIO)